import unittest
import os
import sys
from openpyxl.reader.excel import load_workbook

sys.path.append(os.getcwd())
from sqlite.check_database import check_database
from lib.generate_report import generate_report as gp
from lib.payroll import payroll as p
from lib.oneDriveConnect import oneDriveConnect as odc

"""
This test class tests the generation of the report tally.
"""
class generate_report_test(unittest.TestCase):
    test_db = os.getenv("APPDATA") + "\\project-time-saver\\database_test.db"
    runs = odc.getFiles(path = os.getcwd() + "\\test\\resc\\oct_2021_runs")
    blank_payroll = os.getcwd() + "\\test\\resc\\blank Sheet.xlsx"
    blank_breakdown = os.getcwd() + "\\test\\resc\\Monthly Shift Run Breakdown.xlsx"
    payroll_save = os.getenv("APPDATA") + "\\project-time-saver\\payroll.xlsx"
    breakdown_save = os.getenv("APPDATA") + "\\project-time-saver\\breakdown.xlsx"
    test_config = os.getcwd() + "\\test\\resc\\generate_report_and_payroll_test_config.json"
    test_log = os.getcwd() + "\\test\\resc\\generate_report_test_log.json"

    """
    This method deletes the database, makes a new one, and populates it.
    """
    @classmethod
    def setUpClass(cls):
        generate_report_test.delete_files()
        check_database.check(db_name = "\\database_test.db")
        p.loadWorkBooks(generate_report_test.runs, database = generate_report_test.test_db,
                        test_log_location = generate_report_test.test_log, test_config_location = generate_report_test.test_config)

    """
    This is the actual test.
    """
    def test_generate_reports(self):
        success = gp.generate_report("2021-09-01", "2021-11-15",
                    blank_payroll = generate_report_test.blank_payroll,
                    blank_breakdown = generate_report_test.blank_breakdown,
                    database = generate_report_test.test_db,
                    payroll_save_path = generate_report_test.payroll_save,
                    breakdown_save_path = generate_report_test.breakdown_save)
        self.assertTrue(success)
        self.assertTrue(generate_report_test.check_if_payrolls_are_same(generate_report_test.payroll_save,
                    os.getcwd() + "\\test\\resc\\oct_2021_runs\\expected_reports\\generated_payroll.xlsx"))
        self.assertTrue(generate_report_test.check_breakdowns_are_equal(generate_report_test.breakdown_save,
                    os.getcwd() + "\\test\\resc\\oct_2021_runs\\expected_reports\\generated_monthly_shift_tally.xlsx"))
        generate_report_test.delete_files()

    def check_if_payrolls_are_same(file1, file2):
        """
        Compares each filled cell in the payroll files and determines if they
        are the same.

        inputs...
            file1: the xlsx to read
            file2: the other xlsx to read
        returns...
            true if same false if not
        """
        wb1 = load_workbook(file1)
        wb2 = load_workbook(file2)
        sheet1 = wb1.active
        sheet2 = wb2.active
        notNone = True
        same = True
        row = 8
        while notNone == same == True:
            same = sheet1[f"E{row}"].value == sheet2[f"E{row}"].value
            same = sheet1[f"D{row}"].value == sheet2[f"D{row}"].value
            if sheet1[f"E{row}"].value is None:
                notNone = False
            else:
                row += 1
        if not same: return False
        same = sheet1["E49"].value == sheet2["E49"].value
        if not same: return False
        same = sheet1["E50"].value == sheet2["E50"].value
        if not same: return False
        same = sheet1["E51"].value == sheet2["E51"].value
        if not same: return False
        same = sheet1["E52"].value == sheet2["E52"].value
        if not same: return False
        same = sheet1["E53"].value == sheet2["E53"].value
        if not same: return False
        same = sheet1["E54"].value == sheet2["E54"].value
        if not same: return False
        same = sheet1["E55"].value == sheet2["E55"].value
        if not same: return False
        same = sheet1["E56"].value == sheet2["E56"].value
        if not same: return False
        same = sheet1["E57"].value == sheet2["E57"].value
        return same



    def check_breakdowns_are_equal(file1, file2):
        """
        Compares each filled cell in the breakdown files and determines if they
        are the same.

        inputs...
            file1: the xlsx to read
            file2: the other xlsx to read
        returns...
            true if same false if not
        """
        wb1 = load_workbook(file1)
        wb2 = load_workbook(file2)
        sheet1 = wb1.active
        sheet2 = wb2.active
        same = True
        for i in [4, 5, 6]:
            for letter in ["B", "C", "D", "E", "F", "H", "I"]:
                if sheet1[f"{letter}{i}"].value != sheet2[f"{letter}{i}"].value:
                    return False
                same = sheet1["C8"].value == sheet2["C8"].value
        if not same: return False
        same = sheet1["C9"].value == sheet2["C9"].value
        return same

    """
    Deletes the DB as a part of the setup method.
    """
    def delete_files():
        if os.path.exists(generate_report_test.test_db):
            os.remove(generate_report_test.test_db)
        if os.path.exists(generate_report_test.payroll_save):
            os.remove(generate_report_test.payroll_save)
        if os.path.exists(generate_report_test.breakdown_save):
            os.remove(generate_report_test.breakdown_save)
        if os.path.exists(generate_report_test.test_log):
            os.remove(generate_report_test.test_log)

if __name__ == "__main__":
    unittest.main()