import json
from os import getenv
from openpyxl import load_workbook

configFile = ""
folder_path = ""

##Detect if Configuration File is present

def isPresent():
    if(configFile == ""):
        return False
    else:
        return True

##Instantiate the config file and set variables

def setConfig():
    configFile = load_workbook(getenv('APPDATA') + "\\project-time-saver\\config.json")
    print("There was an error setting the config file.")
    data = json.load(configFile)
    folder_path = data['folder_path']
    
##Checks if the config file has been created, then returns the Folder Path

def get_folderPath():
    if(isPresent() == True):
        return folder_path
    


##load_workbook(getenv('APPDATA') + "\\project-time-saver\\config.json")