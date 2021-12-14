from openpyxl import load_workbook
from .sqlFunctions import sqlFunctions
from .logger import Logger
import os
import traceback
from datetime import datetime

class generate_report:
    # TODO: Change the storage location of the generated files.
    # TODO: Change the path where the blank tally and blank shift
    # breakdowns are stored.
    # TODO: Handle the F.S.C column in the breakdown once we know what it is
    # TODO: Fix bug where top responders are left not being filled correctly
    # TODO: Add error for when no folder is set
    endPaidOnCall = 0
    startFullTime = 0
    endFullTime = 0
    cannotMatch = []
    """
    This function is responsible for being called from the API, running all the
    generation steps, and returning a confirmation or fail message.

    inputs..
        start_date: the first date as a string
        end_date: the last date as a string
    returns..
        case 1: a list containing True in the first position followed be some
            strings with basic details about the report
        case 2: an error message to be displayed to the user
    """
    def generate_report(start_date, end_date, test_log_file = ""):
        generate_report.reset()
        try:
            with sqlFunctions(os.getenv('APPDATA') + "\\project-time-saver\\database.db") as sqlRunner:
                if not sqlRunner.checkForRunsBetweenDates(start_date, end_date):
                    Logger.addNewError("generation error", datetime.now(), "There are no runs for the selected period.", test_log_file)
                else:
                    assert os.path.isfile(os.getenv('APPDATA') + "\\project-time-saver\\blank_tally.xlsx"), "blank_tally.xlsx is missing"
                    assert os.path.isfile(os.getenv('APPDATA') + "\\project-time-saver\\blank_breakdown.xlsx"), "blank_breakdown.xlsx is missing"
                    wb = load_workbook(os.getenv('APPDATA') + "\\project-time-saver\\blank_tally.xlsx")
                    sheet = wb["Sheet1"]
                    number_of_runs = sqlRunner.getNumberOfRuns(start_date, end_date)
                    min_run = sqlRunner.getFirstRunNumber(start_date, end_date)
                    max_run = sqlRunner.getLastRunNumber(start_date, end_date)
                    generate_report.getEndPaidOnCall(sheet)
                    generate_report.getStartFullTime(sheet)
                    generate_report.getEndFullTime(sheet)
                    generate_report.updateEmployeeNulls(sqlRunner, sheet)
                    generate_report.fillTallySheet(sqlRunner, wb, start_date, end_date, min_run, max_run)
                    wb.close()
                    wb = load_workbook(os.getenv('APPDATA') + "\\project-time-saver\\blank_breakdown.xlsx")
                    generate_report.fillBreakdownSheet(sqlRunner, wb, start_date, end_date, min_run, max_run)
                    wb.close()
                    additionalReturns = generate_report.checkForIssues(sqlRunner, min_run, max_run, number_of_runs, start_date, end_date)
                    messages = additionalReturns + [f"The generated pay period is from {start_date} to {end_date}.",
                    f"There were {number_of_runs} runs total this period.", f"This includes runs from run {min_run} to run {max_run}."]
                    for message in messages:
                        Logger.addNewGenerateMessage(message)
        except Exception as e:
            traceback.print_exc()
            print(e)
            Logger.addNewError("generation error", datetime.now(), str(e), test_log_file)

    """
    Resets the global variables. Should not be needed as this class
    uses static references, but for safety this was added.
    """
    def reset():
        generate_report.endPaidOnCall = 0
        generate_report.startFullTime = 0
        generate_report.endFullTime = 0
        generate_report.cannotMatch = []

    """
    This method will do a sanity check on the final product
    to make sure that all of it looks good. If anything is amiss
    (like an employee that could not be matched to the city ID)
    then it will notify the user.

    inputs..
        sqlRunner: a sqlFunctions object
        min_run: the lowest run number
        max_run: the highest run number
        number_of_runs: the total number of runs
        start_date: the starting date of this report
        end_date: the ending date of this report
    returns..
        case 1: an array of strings describing the detected issues
        case 2: None if no sanity issues are found
    """
    def checkForIssues(sqlRunner, min_run, max_run, number_of_runs, start_date, end_date):
        returnArray = ["The report was generated, but failed some sanity checks."]
        missing = generate_report.checkForMissingRuns(sqlRunner, max_run, min_run, number_of_runs, start_date, end_date)
        if missing is not None: returnArray.append(missing)
        cannotBeMatched = generate_report.checkForUnmatchedEmployees(sqlRunner, start_date, end_date)
        if cannotBeMatched is not None: returnArray += cannotBeMatched
        return returnArray if len(returnArray) != 1 else []


    """
    This method checks to ensure that all employees are matched to a city ID.

    inputs..
        sqlRunner: a sqlFunctions object
        start_date: the starting date of this report
        end_date: the ending date of this report
    returns..
        case 1: a list of strings containing a discription of the issue followed
            by a the unmatched employee(s) and some info about them
        case 2: None of no unmatched employees
    """
    def checkForUnmatchedEmployees(sqlRunner, start_date, end_date):
        returnArray = ["One or more employees could not be matched between the run reports\
             and the tally sheet. Ensure all last names are spelled the same on both forms."]
        if generate_report.cannotMatch != []:
            for emp in generate_report.cannotMatch:
                name = emp[1]
                runs = sqlRunner.getCountOfRunsForEmployeeBetweenDates(start_date, end_date, emp[0])
                hours = sqlRunner.getSumOfHoursForEmployeeBetweenGivenDates(start_date, end_date, emp[0])
                returnArray.append(f"{name} could not be matched. They had {runs} run(s) and {hours} hour(s) logged.")
        return returnArray if len(returnArray) != 1 else None

    """
    This method checks for any runs that were not included in this
    report. The is done by counting each run and checking for any
    gaps.

    inputs..
        sqlRunner: a sqlFunctions object
        min_run: the lowest run number
        max_run: the highest run number
        number_of_runs: the total number of runs
        start_date: the starting date of this report
        end_date: the ending date of this report
    returns..
        case 1: a string with the missing run(s)
        case 2: None if no sanity issues are found
    """
    def checkForMissingRuns(sqlRunner, max_run, min_run, number_of_runs, start_date, end_date):
        if max_run - min_run + 1 != number_of_runs:
            numbers = sqlRunner.getOrderedRunsBetweenTwoDates(start_date, end_date)
            missing = []
            for i in range(min_run, max_run + 1):
                if (i,) not in numbers:
                    missing.append(i)
            if len(missing) == 1:
                return f"Run {missing[0]} is missing from the report."
            elif len(missing) > 2:
                returnVal = "Runs "
                for x in missing[:-1]:
                    returnVal += (str(x) + ", ")
                return returnVal + (f"and {missing[-1]} are missing from the report.")
            else:
                return f"Runs {missing[0]} and {missing[1]} are missing from the report."
        else:
            return None


    """
    This method actually fills and saves a copy of the master copy of the tally sheet.

    inputs..
        sqlRunner: a sqlFunctions object
        wb: the xlsx workbook we are working with
        start_date: the first date as a string
        end_date: the last date as a string
    """
    def fillTallySheet(sqlRunner, wb, start_date, end_date, min_run, max_run):
        sheet = wb["Sheet1"]
        for i in range(8, generate_report.endPaidOnCall + 1):
            city_number = sheet[f"A{i}"].value
            if city_number is not None:
                city_number = int(city_number)
                hours = generate_report.getHours(sqlRunner, city_number, start_date, end_date)
                count = generate_report.getCount(sqlRunner, city_number, start_date, end_date)
                sheet[f"D{i}"].value = count
                sheet[f"E{i}"].value = hours
            else:
                sheet[f"D{i}"].value = 0
                sheet[f"E{i}"].value = 0
        for i in range(generate_report.startFullTime, generate_report.endFullTime + 1):
            city_number = sheet[f"A{i}"].value
            if city_number is not None:
                city_number = int(city_number)
                count = generate_report.getCount(sqlRunner, city_number, start_date, end_date)
                sheet[f"D{i}"].value = count
            else:
                sheet[f"D{i}"].value = 0
        sheet["E5"] = min_run
        sheet["G5"] = max_run
        wb.save(os.getenv("APPDATA") + "\\project-time-saver\\tally.xlsx")


    """
    This method is responsible for generating the shift breakdown file.

    inputs..
        sqlRunner: a sqlFunctions object
        wb: the workbook in memory
        start_date: the start date of the period
        end_date: the end date of the period
        min_run: the first run of the period
        max_run: the last run of the period
    """
    def fillBreakdownSheet(sqlRunner, wb, start_date, end_date, min_run, max_run):
        sheet = wb["Sheet1"]
        aWeekdayRuns, bWeekdayRuns, cWeekendRuns = generate_report.getWorkingHourRuns(sqlRunner, start_date, end_date)
        aWeekendRuns, bWeekendRuns, cWeekendRuns = generate_report.getWeekendAndEveningRuns(sqlRunner, start_date, end_date)
        aTotal, bTotal, cTotal = generate_report.getShiftTotals(sqlRunner, start_date, end_date)
        aCover, bCover, cCover = generate_report.getSiftCoverage(sqlRunner, start_date, end_date)
        aStation, bStation, cStation = generate_report.getStationCoverage(sqlRunner, start_date, end_date)
        aMed, bMed, cMed = generate_report.getMedRuns(sqlRunner, start_date, end_date)
        topFT = generate_report.getTopResponder(sqlRunner, start_date, end_date, ft = True)
        topPOC = generate_report.getTopResponder(sqlRunner, start_date, end_date, ft = False)

        sheet["B4"], sheet["B5"], sheet["B6"] = aWeekdayRuns, bWeekdayRuns, cWeekendRuns
        sheet["C4"], sheet["C5"], sheet["C6"] = aWeekendRuns, bWeekendRuns, cWeekendRuns
        sheet["E4"], sheet["E5"], sheet["E6"] = aTotal, bTotal, cTotal
        sheet["F4"], sheet["F5"], sheet["F6"] = aCover, bCover, cCover
        sheet["H4"], sheet["H5"], sheet["H6"] = aStation, bStation, cStation
        sheet["I4"], sheet["I5"], sheet["I6"] = aMed, bMed, cMed
        sheet["C8"], sheet["C9"] = topFT, topPOC
        sheet["A2"] = f"Run {min_run} - Run {max_run}"
        wb.save(os.getenv("APPDATA") + "\\project-time-saver\\breakdown.xlsx")


    """
    This method is responsible for getting the responder who responded to the most runs during the period.

    inputs..
        sqlRunner: a sqlFunctions object
        strt_date: the start of the period
        end_date: the end of the period
        ft: a bool indicating if the top full time employee is wanted or not
    returns..
        A string that contains the top responder(s) followed by the number of runs
    """
    def getTopResponder(sqlRunner, start_date, end_date, ft):
        responders = {}
        results = sqlRunner.getEAndRNumbersFromRespondedBasedOnFullTimeBetweenDates(ft, start_date, end_date)
        filtered_redults = []

        for response in results:
            if sqlRunner.getMedRunBitFromRun(response[1]) == 0:
                filtered_redults.append(response)
        
        for response in filtered_redults:
            if response[0] not in responders:
                responders[response[0]] = 1
            else:
                responders[response[0]] += 1

        largest = 0
        for responder in responders:
            largest = responders[responder] if responders[responder] > largest else largest

        names = []
        for runNumber in [number for number, runs in responders.items() if runs == largest]:
            names.append(sqlRunner.getNameOfEmployeeBasedOnNumber(runNumber))

        if len(names) == 1:
            return f"{names[0]} with {largest} runs"
        elif len(names) == 2:
            return f"{names[0]} and {names[1]} with {largest} runs"
        elif len(names) > 2:
            return f"{names[0]}, " + "".join([f"{name}, " for name in names[1:-1]]) + f"and {names[-1]} with {largest} runs"
        else:
            return "There were no responders in this category for this period."


    """
    This gets the number of med runs for the period.

    inputs..
        sqlRunner: a sqlFunctions object
        strt_date: the start of the period
        end_date: the end of the period
    returns..
        a: the number of med runs for shift a
        b: the number of med runs for shift b
        c: the number of med runs for shift c
    """
    def getMedRuns(sqlRunner, start_date, end_date):
        a = int(sqlRunner.getCountShiftMedRunsBetweenDates("A", start_date, end_date))
        b = int(sqlRunner.getCountShiftMedRunsBetweenDates("B", start_date, end_date))
        c = int(sqlRunner.getCountShiftMedRunsBetweenDates("C", start_date, end_date))
        return a, b, c


    """
    This gets the number of runs where the station was not covered.

    inputs..
        sqlRunner: a sqlFunctions object
        strt_date: the start of the period
        end_date: the end of the period
    returns..
        a: the number of sifts with no station coverage for shift a
        b: the number of sifts with no station coverage for shift b
        c: the number of sifts with no station coverage for shift c
    """
    def getStationCoverage(sqlRunner, start_date, end_date):
        a = int(sqlRunner.getCountShiftNotCoveredRunsBetweenDates("A", start_date, end_date))
        b = int(sqlRunner.getCountShiftNotCoveredRunsBetweenDates("B", start_date, end_date))
        c = int(sqlRunner.getCountShiftNotCoveredRunsBetweenDates("C", start_date, end_date))
        return a, b, c


    """
    Gets the number of runs with 100% shift coverage.

    inputs..
        sqlRunner: a sqlFunctions object
        strt_date: the start of the period
        end_date: the end of the period
    returns..
        a: the number of runs with 100% shift coverage for shift a
        b: the number of runs with 100% shift coverage for shift b
        c: the number of runs with 100% shift coverage for shift c
    """
    def getSiftCoverage(sqlRunner, start_date, end_date):
        a = int(sqlRunner.getCountShiftFullyCoveredRunsBetweenDates("A", start_date, end_date))
        b = int(sqlRunner.getCountShiftFullyCoveredRunsBetweenDates("B", start_date, end_date))
        c = int(sqlRunner.getCountShiftFullyCoveredRunsBetweenDates("C", start_date, end_date))
        return a, b, c


    """
    Gets the total number of fire runs.

    inputs..
        sqlRunner: a sqlFunctions object
        strt_date: the start of the period
        end_date: the end of the period
    returns..
        a: the number of fire runs for shift a
        b: the number of fire runs for shift b
        c: the number of fire runs for shift c
    """
    def getShiftTotals(sqlRunner, start_date, end_date):
        a = int(sqlRunner.getCountShiftFireRunsBetweenDates("A", start_date, end_date))
        b = int(sqlRunner.getCountShiftFireRunsBetweenDates("B", start_date, end_date))
        c = int(sqlRunner.getCountShiftFireRunsBetweenDates("C", start_date, end_date))
        return a, b, c


    """
    Gets the number of runs between 0500 and 1700 M-F.

    inputs..
        sqlRunner: a sqlFunctions object
        strt_date: the start of the period
        end_date: the end of the period
    returns..
        a: the number of working hour runs for shift a
        b: the number of working hour runs for shift b
        c: the number of working hour runs for shift c
    """
    def getWorkingHourRuns(sqlRunner, start_date, end_date):
        a = len(generate_report.stripOffHours(sqlRunner.getDateAndStartOfFireRunsBetweenDatesForShift("A", start_date, end_date)))
        b = len(generate_report.stripOffHours(sqlRunner.getDateAndStartOfFireRunsBetweenDatesForShift("B", start_date, end_date)))
        c = len(generate_report.stripOffHours(sqlRunner.getDateAndStartOfFireRunsBetweenDatesForShift("C", start_date, end_date)))
        return a, b, c

    
    """
    Gets the number of runs between 1700 and 0500 on weekdays and all weekend runs.

    inputs..
        sqlRunner: a sqlFunctions object
        strt_date: the start of the period
        end_date: the end of the period
    returns..
        a: the number of off hour runs for shift a
        b: the number of off hour runs for shift b
        c: the number of off hour runs for shift c
    """
    def getWeekendAndEveningRuns(sqlRunner, start_date, end_date):
        a = len(generate_report.stripWorkingHours(sqlRunner.getDateAndStartOfFireRunsBetweenDatesForShift("A", start_date, end_date)))
        b = len(generate_report.stripWorkingHours(sqlRunner.getDateAndStartOfFireRunsBetweenDatesForShift("B", start_date, end_date)))
        c = len(generate_report.stripWorkingHours(sqlRunner.getDateAndStartOfFireRunsBetweenDatesForShift("C", start_date, end_date)))
        return a, b, c


    """
    Takes a list of dates and times and returns a list of dates 
    and times that are on the weekend or between 1700 and 0500.
    """
    def stripWorkingHours(toStrip):
        return [run for run in toStrip if not generate_report.isWorkingHours(run)]

    
    """
    Takes a list of dates and times and returns a list of dates 
    and times that are during weekdays between 0500 and 1700.
    """
    def stripOffHours(toStrip):
        return [run for run in toStrip if generate_report.isWorkingHours(run)]


    """
    Determines if a date is on a weekday between 0500 and 1700 or not.

    inputs..
        date_time: a tuple with a string date in position 0 and an
            int that represents a military time in position 1
    returns..
        case 1: True if it is a weekday between 0500 and 1700
        case 2: False if not
    """
    def isWorkingHours(date_time):
        return True if datetime.strptime(date_time[0], "%Y-%m-%d").weekday() < 5 \
            and (date_time[1] >= 500 and date_time[1] <= 1700) else False


    """
    This method gets the number of runs a specific person
    responded to in a given period.

    inputs..
        sqlRunner: a sqlFunctions object
        city_number: the city assigned employee ID
        start_date: the first date as a string
        end_date: the last date as a string
    returns..
        case 1: the number of runs in a given period
    """
    def getCount(sqlRunner, city_number, start_date, end_date):
        total = 0
        runNumbers = sqlRunner.getAllRunsNumbersEmployeeByCityNumberRespondedToBetweenDates(start_date, end_date, city_number)
        for run in runNumbers:
            total += 1 if sqlRunner.getMedRunBitFromRun(run[0]) == 0 else 0
        return total

    """
    This method gets the number of hours a specific person
    worked on runs in a given period.

    inputs..
        sqlRunner: a sqlFunctions object
        city_number: the city assigned employee ID
        start_date: the first date as a string
        end_date: the last date as a string
    returns..
        case 1: the number of hours a given employee worked
    """
    def getHours(sqlRunner, city_number, start_date, end_date):
        total = 0
        runs = sqlRunner.getRunNumberOfAllPaidRunsForEmplyeeByEmployeeNumberBetweenDates(city_number, start_date, end_date)
        for run in runs:
            hour = sqlRunner.getRunTimeOfFireRunByRunNumber(run[0])
            if len(hour) == 1:
                hour = hour[0][0]
            else:
                hour = 0
            if hour is not None and hour != []:
                total += hour
        return total

    """
    This method updates the Employee table to ensure no employees have a NULL
    value in the city_number column. If they do they will not be paid.

    inputs..
        sqlRunner: a sqlFunctions object
        sheet: the sheet for the tally xlsx
    """
    def updateEmployeeNulls(sqlRunner, sheet):
        nullEmps = sqlRunner.getNumberAndNameOfAllEmployeesWithNoCityNumver()
        generate_report.insertCityIDs(sqlRunner, nullEmps, sheet)

    """
    This method actually inserts the city IDs for update_employee_nulls

    inputs..
        conn: the connection to the SQL
        nullEmps: a list of employee rows that have null city_number values 
        sheet: the sheet for the tally xlsx
    """
    def insertCityIDs(sqlRunner, nullEmps, sheet):
        for emp in nullEmps:
            unmatched = emp
            for i in range(8, generate_report.endFullTime + 1):
                if sheet[f"A{i}"].value is not None and \
                        generate_report.match_names(emp[1], sheet[f"B{i}"].value, sheet[f"C{i}"].value):
                    sqlRunner.addCityNumberToEmployee(sheet[f"A{i}"].value, emp[0])
                    unmatched = None
                    break
            if unmatched is not None:
                generate_report.cannotMatch.append(emp)
                

    """
    This method gets the final row paid on call employees.

    inputs..
        sheet: the sheet for the tally xlsx
    """
    def getEndPaidOnCall(sheet):
        end = False
        if generate_report.endPaidOnCall == 0:
            generate_report.endPaidOnCall = 8
            while (not end):
                if sheet[f"C{generate_report.endPaidOnCall + 1}"].value != None:
                    generate_report.endPaidOnCall = generate_report.endPaidOnCall + 1
                else:
                    end = True

    """
    This method gets the first row of full time employees.

    inputs..
        sheet: the sheet for the tally xlsx
    """
    def getStartFullTime(sheet):
        end = False
        if generate_report.startFullTime == 0:
            generate_report.startFullTime = generate_report.endPaidOnCall
            while (not end):
                if sheet[f"A{generate_report.startFullTime + 1}"].value == None:
                    generate_report.startFullTime = generate_report.startFullTime + 1
                else:
                    generate_report.startFullTime = generate_report.startFullTime + 1
                    end = True

    """
    This method gets the final row that we are concerned with editing.

    inputs..
        sheet: the sheet for the tally xlsx
    """
    def getEndFullTime(sheet):
        end = False
        if generate_report.endFullTime == 0:
            generate_report.endFullTime = generate_report.startFullTime
            while (not end):
                if sheet[f"A{generate_report.endFullTime + 1}"].value != None:
                    generate_report.endFullTime = generate_report.endFullTime + 1
                else:
                    end = True

    """
    This method will take a name from the run reports and compare it to
    a name in the tally xlsx in order to determine if they are the same
    person.

    inputs..
        name: the name from the run report/responded table
        fname: the first name from the tally
        lname: the last name from the tally
    returns..
        case 1: True if they are the same person
        case 2: False if they aren't the same person
    """
    def match_names(name, fname, lname):
        if name[0:4] == "Lt. ":
             name = name[4:]
        elif name[0:3] == "Lt.":
            name = name[3:]
        elif name[0:6] == "Capt. ":
            name = name[6:]
        elif name[0:5] == "Capt.":
            name = name[5:]
        if name[-1].isnumeric():
            if name[-2].isnumeric():
                if name[-5] != "-":
                    name = name[0:-4]
                else:
                    name = name[0:-6]
            else:
                if name[-4] != "-":
                    name = name[0:-3]
                else:
                    name = name[0:-5]

        if name == f"{fname[0]}. {lname}" \
                or name == f"{fname[0]}.{lname}":
            return True
        else:
             return False